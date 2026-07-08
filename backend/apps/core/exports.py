import csv
import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle

# Emboita Hotel brand gold — mirrors frontend/src/index.css --color-brand-*
BRAND_GOLD = "#f2a900"
BRAND_GOLD_DARK = "#a56700"
BRAND_GOLD_LIGHT = "#fffbea"
BRAND_TEXT_DARK = "#1f2937"


def _company_profile():
    from apps.core.models import CompanyProfile

    return CompanyProfile.get_solo()


def _report_title(filename):
    return filename.replace(".csv", "").replace(".xlsx", "").replace(".pdf", "").replace("_", " ").title()


def export_csv(rows, headers, filename="export.csv"):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def export_xlsx(rows, headers, filename="export.xlsx"):
    profile = _company_profile()
    workbook = Workbook()
    sheet = workbook.active

    header_fill = PatternFill(start_color=BRAND_GOLD.lstrip("#"), end_color=BRAND_GOLD.lstrip("#"), fill_type="solid")
    zebra_fill = PatternFill(start_color=BRAND_GOLD_LIGHT.lstrip("#"), end_color=BRAND_GOLD_LIGHT.lstrip("#"), fill_type="solid")
    dark_font = Font(color=BRAND_TEXT_DARK.lstrip("#"), bold=True)

    title_row = 1
    header_row = 2
    if profile.logo:
        try:
            img = XlsxImage(profile.logo.path)
            img.height, img.width = 42, 140
            sheet.add_image(img, "A1")
            title_row = 4
            header_row = 5
        except FileNotFoundError:
            pass

    sheet.cell(row=title_row, column=1, value=f"{profile.name} — {_report_title(filename)}").font = Font(
        bold=True, size=13, color=BRAND_TEXT_DARK.lstrip("#")
    )

    for col, heading in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=heading)
        cell.fill = header_fill
        cell.font = dark_font
        cell.alignment = Alignment(horizontal="left")

    for r_offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            cell = sheet.cell(row=header_row + 1 + r_offset, column=col, value=value)
            if r_offset % 2 == 1:
                cell.fill = zebra_fill

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


ATTENDANCE_STATUS_FILL = {
    "P": "C6EFCE",
    "L": "FFEB9C",
    "A": "FFC7CE",
    "LV": "BDD7EE",
    "OFF": "F2F2F2",
}
ATTENDANCE_STATUS_TEXT = {
    "P": "006100",
    "L": "9C6500",
    "A": "9C0006",
    "LV": "1F4E78",
    "OFF": "808080",
}


def export_attendance_grid_xlsx(rows, headers, status_col_start, filename="attendance_grid.xlsx"):
    """Like export_xlsx, but colors each attendance-status cell (P/L/A/LV/OFF)
    and freezes the info columns + header row so a wide date range stays
    readable while scrolling — used for the Employee/Department/Company
    present-absent grid report, not the generic flat-row exports.
    """
    profile = _company_profile()
    workbook = Workbook()
    sheet = workbook.active

    header_fill = PatternFill(start_color=BRAND_GOLD.lstrip("#"), end_color=BRAND_GOLD.lstrip("#"), fill_type="solid")
    zebra_fill = PatternFill(start_color=BRAND_GOLD_LIGHT.lstrip("#"), end_color=BRAND_GOLD_LIGHT.lstrip("#"), fill_type="solid")
    dark_font = Font(color=BRAND_TEXT_DARK.lstrip("#"), bold=True)

    title_row = 1
    header_row = 2
    if profile.logo:
        try:
            img = XlsxImage(profile.logo.path)
            img.height, img.width = 42, 140
            sheet.add_image(img, "A1")
            title_row = 4
            header_row = 5
        except FileNotFoundError:
            pass

    sheet.cell(row=title_row, column=1, value=f"{profile.name} — {_report_title(filename)}").font = Font(
        bold=True, size=13, color=BRAND_TEXT_DARK.lstrip("#")
    )

    for col, heading in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=heading)
        cell.fill = header_fill
        cell.font = dark_font
        cell.alignment = Alignment(horizontal="center" if col > status_col_start else "left")

    for r_offset, row in enumerate(rows):
        for col, value in enumerate(row, start=1):
            cell = sheet.cell(row=header_row + 1 + r_offset, column=col, value=value)
            if col > status_col_start:
                fill = ATTENDANCE_STATUS_FILL.get(value)
                if fill:
                    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
                    cell.font = Font(color=ATTENDANCE_STATUS_TEXT.get(value, BRAND_TEXT_DARK.lstrip("#")), bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif r_offset % 2 == 1:
                cell.fill = zebra_fill

    for col in range(1, status_col_start + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20
    for col in range(status_col_start + 1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 7

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=status_col_start + 1)

    legend_row = header_row + len(rows) + 2
    legend = "Legend:  P = Present   L = Late   A = Absent   LV = On Leave   OFF = Non-working day"
    sheet.cell(row=legend_row, column=1, value=legend).font = Font(italic=True, size=9, color="64748b")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _format_cell(value):
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _column_widths(headers, rows, available_width):
    """Proportional column widths (by sampled content length) that always sum to
    exactly `available_width`, so the table spans the page without ever overflowing."""
    sample = rows[:50]
    weights = []
    for i in range(len(headers)):
        header_len = len(str(headers[i])) + 2  # bias so short headers don't wrap
        cell_lens = [len(_format_cell(r[i])) for r in sample if i < len(r)]
        longest = max([header_len, *cell_lens]) if cell_lens else header_len
        weights.append(min(max(longest, 7), 45))
    total_weight = sum(weights)
    return [available_width * (w / total_weight) for w in weights]


def export_pdf(rows, headers, filename="export.pdf"):
    profile = _company_profile()
    buffer = io.BytesIO()
    margin = 14 * mm
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=margin,
        bottomMargin=margin,
        leftMargin=margin,
        rightMargin=margin,
    )

    flowables = []
    title_style = ParagraphStyle("ReportTitle", fontSize=15, textColor=colors.HexColor(BRAND_TEXT_DARK), spaceAfter=2)
    subtitle_style = ParagraphStyle("ReportSubtitle", fontSize=10, textColor=colors.HexColor(BRAND_GOLD_DARK))

    if profile.logo:
        try:
            logo = PdfImage(profile.logo.path, width=32 * mm, height=9 * mm, kind="proportional")
            logo.hAlign = "LEFT"
            flowables.append(logo)
            flowables.append(Spacer(1, 8))
        except (FileNotFoundError, OSError):
            pass

    flowables.append(Paragraph(profile.name or "Emboita Hotel", title_style))
    flowables.append(Paragraph(_report_title(filename), subtitle_style))
    flowables.append(Spacer(1, 10))

    num_cols = len(headers)
    font_size = 8 if num_cols <= 7 else 7 if num_cols <= 10 else 6
    header_style = ParagraphStyle(
        "TableHeader", fontName="Helvetica-Bold", fontSize=font_size, leading=font_size + 2,
        textColor=colors.HexColor(BRAND_TEXT_DARK),
    )
    body_style = ParagraphStyle(
        "TableBody", fontName="Helvetica", fontSize=font_size, leading=font_size + 2,
        textColor=colors.HexColor(BRAND_TEXT_DARK),
    )

    data = [[Paragraph(str(h), header_style) for h in headers]]
    for row in rows:
        data.append([Paragraph(_format_cell(cell), body_style) for cell in row])

    col_widths = _column_widths(headers, rows, doc.width)
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_GOLD)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5c98a")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(BRAND_GOLD_LIGHT)]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    flowables.append(table)
    doc.build(flowables)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_queryset(request, queryset, headers, row_fn, base_filename):
    """Renders a queryset to csv/xlsx/pdf based on `?format=`. Returns None if unsupported."""
    fmt = request.query_params.get("format", "").lower()
    rows = [row_fn(obj) for obj in queryset]
    if fmt == "csv":
        return export_csv(rows, headers, f"{base_filename}.csv")
    if fmt == "xlsx":
        return export_xlsx(rows, headers, f"{base_filename}.xlsx")
    if fmt == "pdf":
        return export_pdf(rows, headers, f"{base_filename}.pdf")
    return None
